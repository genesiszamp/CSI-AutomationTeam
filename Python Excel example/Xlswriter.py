import xlsxwriter
from openpyxl import Workbook, load_workbook

wb = load_workbook('SLMS.xlsx')
ws = wb.active



for row in ws.iter_rows():
    list = []
    list.append(row[1].value)
    list.remove("Overdue Trainings")
    list = [int(i) for i in list]
    Overdue_Trainings = sum(list)


for row_num, data in enumerate(list):
    workbook = xlsxwriter.workbook('SLMS.xlsx')
    worksheet = workbook.add_worksheet('Sum')
    worksheet.write(row_num, 0, data)
    worksheet.write(row_num, 1, ws.title)

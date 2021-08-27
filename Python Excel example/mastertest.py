from openpyxl import Workbook, load_workbook

wb = load_workbook('SLMS.xlsx')
ws = wb.active
sheet_names = ws.title
write = wb.create_sheet('Sum',0)

def main():
    for sheet in wb.worksheets:
        data()
        
        
def data():
    for row in ws.iter_rows():
        list = []
        list.append(row[1].value)
        if 'Overdue Trainings' in list:
            list.remove("Overdue Trainings")
            print("Si elimino la shit")
        else:
            list = [int(i) for i in list]       
            Overdue_Trainings = sum(list)
            summary(Overdue_Trainings)
            print("Si paso la lista")

def summary(Overdue_Trainings):
    
    write.append([ws.title,Overdue_Trainings])
    wb.save("SLMS.xlsx")
    print("Printeo la jugada")

  

if __name__ == "__main__":
    main()
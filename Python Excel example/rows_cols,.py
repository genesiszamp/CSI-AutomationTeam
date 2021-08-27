from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Grades.xlsx')
ws = wb['Grades']

for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(row)
        print(ws[char + str(col)].value)
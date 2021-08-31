from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
 
wb = load_workbook('SLMS.xlsx')

ws = wb
sheets = wb.sheetnames
n = 0
i=0
list = []
total = []



for ws in wb.worksheets:
     #Move along worksheets
     ws = wb[sheets[n]]
     
      #Create a list, convert to int  
     for row in ws.iter_rows(min_row=2,min_col=1):
        list.append(int(row[1].value))
        
        
     total.append(sum(list))
     list.clear()
     n = n + 1
total.append(sum(total))

write = wb.create_sheet('Total')

write['A1'] = 'Managers'
write['B1'] = 'Overdue Trainings'
write['A1'].font = Font(bold=True)
write['B1'].font = Font(bold=True)


for ws in wb.worksheets: 
   write.append([ws.title,total[i]])
   i = i + 1        
wb.save("SLMS.xlsx")


        
print(total)